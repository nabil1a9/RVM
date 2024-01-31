from django.shortcuts import render , redirect
from django.http import HttpResponse , HttpResponseRedirect , JsonResponse,HttpRequest
from .models import Boutaille , Note , User
from .forms import RateForm
from django.views.decorators.csrf import csrf_exempt
import requests
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
import os
import time
from openpyxl import load_workbook


# Create your views here.
def ajouter(response):


    wb = load_workbook(filename='C:/Users/Public/yolov71/data.xlsx', read_only=True)

        # Get the active worksheet
    ws = wb.active

        # Get the last row number of the worksheet
    last_row = ws.max_row

        # Get the last row of the worksheet as a list
    last_row_data = [cell.value for cell in ws[last_row]]

        # Assign each cell to a variable
    classe = last_row_data[0]
    etat = last_row_data[1]
    cap = last_row_data[2]
    marque = last_row_data[3]
        # and so on...

        # Do something with the variables. For example, you could print them:
    b = response.user.boutaille.create(type="ALUMINIUM",etat="proper",cap="nocap",marque="fanta")
    return HttpResponseRedirect("/afficher")


def calcul():
    a = Note.objects.all()

    total = 0
    if(Note.objects.count()>0):
        for r in a:
            total += r.rate
        return total / Note.objects.count()
    else:
        return 0

def index(response,id):
    liste = Boutaille.objects.get(id=id)
    return render(response,"main/index.html",{})


def afficher(response):
    # Initialize last modification time of xlsx file

    a = calcul()
    i=0
    j=0

    b=Boutaille.objects.filter(user = response.user)
    for r in b :
        if (r.type== "PLASTIC") :
            i=i+1
        else :
            j=j+1
    if response.method == "POST":
        form = RateForm(response.POST)
        if form.is_valid():
            n = form.cleaned_data["rate"]
            t = Note(rate=n)
            t.save()


    else:
        form=RateForm()

    return render(response,"main/index.html",{"b":b,"i":i,"j":j,"form":form,"moy":a})



def rate( rating: int) :

    n=Note(rate=rating)
    n.save()
    return HttpResponseRedirect("/afficher")

def process_data(request):
    if request.method == 'POST':
        my_variable = request.POST.get('result')
        print(my_variable)  # print value of my_variable
    else:
        my_variable = None
    context = {'my_variable': my_variable}
    return render(request, 'main/json.html', context)

